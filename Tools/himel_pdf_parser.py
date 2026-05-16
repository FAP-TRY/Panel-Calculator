#!/usr/bin/env python3
"""
himel_pdf_parser.py
===================
Reads Himel 2025 price list PDF and exports to CSV + Excel
with format suitable for Panel Calculator database import.

Output columns:
  category, reference_code, product_name, specifications,
  price, stock_status, vendor, price_year

Discount: NONE — harga list asli dari PDF (tanpa diskon diterapkan).
Catatan: diskon 45%+10% sudah dihapus per 2026-05-16 sesuai permintaan.

Stock Status: 1 = Stock (default)

Usage:
  python himel_pdf_parser.py [path_to_pdf] [price_year]
"""

import re
import sys
import csv
import os
from pathlib import Path

# Force UTF-8 output on Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Dependency check ──────────────────────────────────────────────────────────
try:
    import pdfplumber
except ImportError:
    print("[INFO] Installing pdfplumber...")
    os.system("pip install pdfplumber")
    import pdfplumber

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# ── Constants ─────────────────────────────────────────────────────────────────
VENDOR = "Himel"
DEFAULT_PRICE_YEAR = 2025
STOCK_STATUS = 1
DISCOUNT_FACTOR = 1.0  # No discount — use raw list price from PDF

# ── Regex ─────────────────────────────────────────────────────────────────────

# Price: Indonesian format at end of line, e.g. 1.234.567
PRICE_RE = re.compile(r'(\d{1,3}(?:\.\d{3})+)\s*$')

# Himel reference codes: start with H + uppercase letters + digits
# e.g. HAVBR50, HAPSR1AC12, HKGG316T230, HTND5HI100EWF, HAVHFL0022KWA10
# Also covers FIBOX (accessory), HEBSP, HWDA, HWDS, HHEF, HHE, etc.
HIMEL_REF_RE = re.compile(
    r'\b(H[A-Z]{1,4}[A-Z0-9]{3,}|FIBOX\w+)\b'
)

# cid garbage: (cid:xx) sequences
CID_RE = re.compile(r'\(cid:\d+\)')

# Bullet / special chars at start of line
BULLET_RE = re.compile(r'^[^\w\s(]+')


def clean_cid(text):
    """Remove (cid:xx) encoding garbage from text."""
    return CID_RE.sub('', text).strip()


def clean_price(price_str):
    """Convert Indonesian price string '1.234.567' to integer."""
    return int(price_str.replace('.', ''))


def apply_discount(list_price):
    """Apply 45% + 10% compound discount."""
    return round(list_price * DISCOUNT_FACTOR)


# ── Category detection ────────────────────────────────────────────────────────

# Ordered list of (pattern, category) — first match wins
CATEGORY_PATTERNS = [
    # Motor Management subcategories
    (re.compile(r'Breaking Unit|Braking Unit', re.I), 'Motor Management - Braking Unit VSD'),
    (re.compile(r'DV/DT Filter|DVDT Filter', re.I),  'Motor Management - VSD DV/DT Filter'),
    (re.compile(r'Passive Filter', re.I),             'Motor Management - VSD Passive Filter'),
    (re.compile(r'Smart Relay',    re.I),             'Motor Management - Smart Relay'),
    (re.compile(r'\bPLC\b',        re.I),             'Motor Management - PLC'),
    # Voltage Stabilizer
    (re.compile(r'Automatic Voltage Regulator|Voltage Regulator|AVR', re.I), 'Voltage Stabilizer - AVR'),
    # Control Components
    (re.compile(r'Digital Time Switch|Time Switch', re.I), 'Control Components - Digital Time Switch'),
    # Home Electric subcategories
    (re.compile(r'EV Charg',     re.I),  'Home Electric - EV Charging'),
    (re.compile(r'Tools.*Meter|Meter.*Tool|ELECTRICAL TAPE', re.I), 'Home Electric - Tools & Meters'),
    (re.compile(r'Style Series', re.I),  'Home Electric - Style Series Wiring Devices'),
    (re.compile(r'Curvo Series', re.I),  'Home Electric - Curvo Series Wiring Devices'),
    (re.compile(r'WD Box',       re.I),  'Home Electric - WD Box'),
    (re.compile(r'Floor Socket', re.I),  'Home Electric - Floor Socket'),
    (re.compile(r'Wiring Device',re.I),  'Home Electric - Wiring Devices'),
    (re.compile(r'HOME ELECTRIC', re.I), 'Home Electric'),
    (re.compile(r'MOTOR MANAGEMENT', re.I), 'Motor Management'),
    (re.compile(r'VOLTAGE STABILIZER', re.I), 'Voltage Stabilizer'),
    (re.compile(r'CONTROL COMPONENTS', re.I), 'Control Components'),
]


def determine_category(page_header_text):
    """Determine category from page text (uses first 400 chars, then full page)."""
    # Try first 400 chars for most pages
    for pattern, cat in CATEGORY_PATTERNS:
        if pattern.search(page_header_text):
            return cat
    return 'Other'


# ── Product name builders per product family ──────────────────────────────────

def build_product_name(ref_code, after_ref_raw, subcategory):
    """
    Build a clean product_name from the text after the reference code.
    For module-type products (Smart Relay, PLC, VSD filters), derive name
    from the reference code since after_ref contains table column data.
    """
    # For module/component types, always use derived name (after_ref is specs)
    if ref_code.startswith(('HAPSR', 'HAPSRE', 'HAPSM', 'HAP226', 'HAPAMS', 'HAP291', 'HAHG')):
        return derive_name_from_ref(ref_code, subcategory)
    if ref_code.startswith(('HAVVTR', 'HAVHFL', 'HAVBR')):
        return derive_name_from_ref(ref_code, subcategory)
    if ref_code.startswith('HKG'):
        return derive_name_from_ref(ref_code, subcategory)
    if ref_code.startswith('HHEV'):
        return derive_name_from_ref(ref_code, subcategory)

    # For descriptive products (EV chargers, wiring devices, floor sockets, etc.)
    # the after_ref text usually IS the description
    text = clean_cid(after_ref_raw)
    text = HIMEL_REF_RE.sub('', text)
    text = re.sub(r'[^\w\s().,/+%-]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    # Remove leading pure-digit tokens (standalone numbers like "0,75" "5" "2,2 7")
    # but keep alphanumeric like "13A", "7KW"
    text = re.sub(r'^\s*\d[\d,.\s]+(?=[A-Za-z])', '', text).strip()  # "2,2 7 " before word
    text = re.sub(r'^\s*\d+\s+(?=\d)', '', text).strip()             # leading lone digit
    text = re.sub(r'^[^\w(]+|[^\w)]+$', '', text).strip()

    if len(text) >= 3:
        return text

    return derive_name_from_ref(ref_code, subcategory)


def derive_name_from_ref(ref_code, subcategory):
    """Derive a human-readable product name from the reference code."""
    # VSD DV/DT filters
    if ref_code.startswith('HAVVTR'):
        m = re.search(r'(\d+)', ref_code[6:])
        ampere = m.group(1).lstrip('0') or '0' if m else ''
        return f"VSD DV/DT Filter {ampere}A"

    # VSD Passive filters
    if ref_code.startswith('HAVHFL'):
        m = re.search(r'(\d+)KWA', ref_code)
        kw = (m.group(1).lstrip('0') or '0') if m else ''
        return f"VSD Passive Filter {kw}kW"

    # Braking units
    if ref_code.startswith('HAVBR'):
        amps = ref_code[5:].lstrip('0') or '0'
        return f"Braking Unit {amps}A"

    # Voltage stabilizers
    if ref_code.startswith('HTND') or ref_code.startswith('HSJW') or ref_code.startswith('HSBW'):
        return "Automatic Voltage Regulator"

    # Smart relay / PLC
    if ref_code.startswith('HAPSREAM'):
        return "Smart Relay Analog Extension Module"
    if ref_code.startswith('HAPSREDM') or ref_code.startswith('HAPSREMIO'):
        return "Smart Relay Digital Extension Module"
    if ref_code.startswith('HAPSR2'):
        return "Smart Relay with Ethernet"
    if ref_code.startswith('HAPSR'):
        return "Smart Relay Main Module"
    if ref_code.startswith('HAPSM'):
        return "PLC Signal Module"
    if ref_code.startswith('HAPSM2317'):
        return "PLC Thermocouple/RTD Module"
    if ref_code.startswith('HAP226'):
        return "PLC CPU Module"
    if ref_code.startswith('HAPAMS'):
        return "PLC Expansion Board"
    if ref_code.startswith('HAP291'):
        return "PLC Memory Card"
    if ref_code.startswith('HAHG'):
        # Extract screen size from ref: HAHG070E = 7-inch, HAHG101E = 10.1-inch
        m = re.search(r'(\d+)', ref_code[4:])
        if m:
            size_raw = m.group(1).lstrip('0') or '0'
            size = int(size_raw)
            # 70 → 7.0 → "7", 101 → 10.1
            if size >= 100:
                display_size = f"{size/10:.1f}".rstrip('0').rstrip('.')
            else:
                # e.g. 70 → should be 7.0 inch (leading zero stripped: 070 → 70)
                display_size = f"{size/10:.1f}".rstrip('0').rstrip('.')
            return f"HMI Touch Screen {display_size}-inch"
        return "HMI Touch Screen"

    # Digital time switches
    if ref_code.startswith('HKG'):
        return "Digital Time Switch"

    # EV chargers
    if ref_code.startswith('HHEV'):
        # Detect AC vs DC and power from reference code
        if '2107HP' in ref_code:
            m = re.search(r'HP(\d+)K', ref_code)
            kw = m.group(1) if m else ''
            return f"AC EV Charger {kw}kW" if kw else "AC EV Charger"
        elif '3401HM' in ref_code:
            m = re.search(r'(\d+)K', ref_code)
            kw = m.group(1) if m else ''
            return f"DC EV Charger {kw}kW" if kw else "DC EV Charger"
        elif '3701HM' in ref_code:
            m = re.search(r'(\d+)K', ref_code)
            kw = m.group(1) if m else ''
            return f"DC EV Charger {kw}kW" if kw else "DC EV Charger"
        return "EV Charger"

    # Home electric wiring devices
    if ref_code.startswith('HWDS') or ref_code.startswith('HWDA') or ref_code.startswith('HWDI') or ref_code.startswith('HWDO'):
        return "Wiring Device"

    if ref_code.startswith('HWDG') or ref_code.startswith('HWDIG') or ref_code.startswith('HWDIR'):
        return "WD Box"

    if ref_code.startswith('HHEF'):
        return "Floor Socket"

    if ref_code.startswith('HHE'):
        return "Electrical Tool/Meter"

    if ref_code.startswith('FIBOX'):
        return "Inbow Box"

    if ref_code.startswith('HEBSP'):
        return "BS Plug"

    # Generic fallback
    return ref_code


def build_specifications(before_ref_raw, after_ref_raw):
    """
    Build specifications string from context around the reference code.
    Uses 'before_ref' (e.g. power/ampere columns) and post-ref technical data.
    """
    # Combine before and after
    parts = []

    before = clean_cid(before_ref_raw)
    # Strip other ref codes
    before = HIMEL_REF_RE.sub('', before)
    before = re.sub(r'[^\w\s().,/+%-]', ' ', before)
    before = re.sub(r'\s+', ' ', before).strip()
    # Keep numeric specs from before (e.g. "2,2 7" = 2.2kW 7A)
    if before and re.search(r'\d', before):
        parts.append(before)

    after = clean_cid(after_ref_raw)
    after = HIMEL_REF_RE.sub('', after)
    after = re.sub(r'[^\w\s().,/+%-]', ' ', after)
    after = re.sub(r'\s+', ' ', after).strip()
    after = re.sub(r'^[^\w(]+|[^\w)]+$', '', after).strip()
    if after and len(after) >= 3:
        parts.append(after)

    spec = ' | '.join(p for p in parts if p)
    # Limit length
    if len(spec) > 150:
        spec = spec[:147] + '...'
    return spec


# ── Per-page extraction ───────────────────────────────────────────────────────

def merge_split_lines(lines):
    """
    Some product rows are split across two lines in the PDF, e.g.:
      "1 Gang Schuko Socket"
      "HWDSSS 28.888"
    Merge such continuation lines: if a line has no price and the next has
    a Himel ref + price, join them.
    """
    merged = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue

        # Check if next line starts with a Himel ref followed by a price
        if i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            next_has_ref = HIMEL_REF_RE.match(next_line)
            next_has_price = PRICE_RE.search(next_line)
            this_has_price = PRICE_RE.search(line)
            this_is_header = re.search(r'\bPicture\b', line) and re.search(r'\b(Reference|Decription|Description)\b', line)

            this_has_ref = bool(HIMEL_REF_RE.search(line))
            # Also skip merging if this line looks like a table column header continuation,
            # section label, or technical abbreviation
            # e.g. "Voltage or AI) Input Output Output tion (Rp)"
            # or "1xRS485,1xRJ45" or "(Ethernet)" or "MAIN MODULE" or "EXTENSION MODULE"
            this_is_col_header = bool(
                (re.search(r'\b(Rp|Input|Output|Voltage|Communication|Display|Ethernet|RS485|RJ45|MODULE|SECTION)\b', line, re.I)
                 and len(line) > 3)
                or re.match(r'^[\d\w,.()/\s]+$', line) and not re.search(r'[a-zA-Z]{4,}', line)
            )
            if next_has_ref and next_has_price and not this_has_price and not this_is_header and not this_has_ref and not this_is_col_header:
                # Join: pure-text description on this line + ref+price on next
                merged.append(line + ' ' + next_line)
                i += 2
                continue

        merged.append(line)
        i += 1
    return merged


def extract_products_from_page(page, category):
    """
    Extract product rows from a PDF page using line-by-line analysis.
    Handles two layouts:
      A) "REFCODE [specs] PRICE"  (most pages)
      B) "Description REFCODE PRICE"  (style/wiring device pages, page 11)
    """
    text = page.extract_text()
    if not text:
        return []

    raw_lines = text.split('\n')
    lines = merge_split_lines(raw_lines)
    products = []

    for line in lines:
        original_line = line
        line = line.strip()
        if not line:
            continue

        # Must have a price at end
        price_match = PRICE_RE.search(line)
        if not price_match:
            continue

        price_str = price_match.group(1)
        price_end = price_match.start()
        line_before_price = line[:price_end].strip()

        list_price = clean_price(price_str)

        # Skip obvious non-product prices (page numbers, ZIP codes, etc.)
        if list_price < 1000:
            continue

        # Skip table header rows (lines containing "Picture" + "Reference"/"Decription")
        if re.search(r'\bPicture\b', line) and re.search(r'\b(Reference|Decription|Description)\b', line):
            continue

        # Skip "applicable for" compatibility reference codes that appear in
        # VSD filter tables: HAVBA4T*, HAVXS4T*, HAVSP4T*
        # These are not standalone products in the pricelist — they're cross-refs
        if re.match(r'(HAVBA4T|HAVXS4T|HAVSP4T)', line.strip()):
            continue

        # Find the FIRST Himel reference code in this line
        ref_match = HIMEL_REF_RE.search(line_before_price)
        if not ref_match:
            continue

        ref_code = ref_match.group(1)
        ref_start = ref_match.start()
        ref_end = ref_match.end()

        # Text before reference code (could be power/ampere column data)
        before_ref = line_before_price[:ref_start].strip()
        # Text after reference code (up to price)
        after_ref = line_before_price[ref_end:].strip()

        # Detect layout:
        # Layout A: REFCODE is first token → description comes after ref
        # Layout B: REFCODE is after description text → before_ref has description
        is_layout_b = len(before_ref) > 5 and re.search(r'[a-zA-Z]{3,}', before_ref)

        if is_layout_b:
            # Description is before_ref; after_ref may have extra ref codes (applicable for)
            # BUT for module types, always use derived name even in layout B
            # (the before_ref is table column data, not a real description)
            module_prefixes = ('HAPSR', 'HAPSRE', 'HAPSM', 'HAP226', 'HAPAMS', 'HAP291', 'HAHG',
                               'HAVVTR', 'HAVHFL', 'HAVBR', 'HHEV')
            if any(ref_code.startswith(p) for p in module_prefixes):
                product_name = derive_name_from_ref(ref_code, category)
                specifications = build_specifications('', after_ref)
            else:
                desc_clean = clean_cid(before_ref)
                desc_clean = HIMEL_REF_RE.sub('', desc_clean).strip()
                desc_clean = re.sub(r'[^\w\s().,/+%-]', ' ', desc_clean)
                desc_clean = re.sub(r'\s+', ' ', desc_clean).strip()
                desc_clean = re.sub(r'^[^\w]+|[^\w)]+$', '', desc_clean).strip()

                if len(desc_clean) >= 3:
                    product_name = desc_clean
                else:
                    product_name = derive_name_from_ref(ref_code, category)
                specifications = build_specifications('', after_ref)
        else:
            # Layout A: ref is first — derive product name from ref, use after_ref as specs
            product_name = build_product_name(ref_code, after_ref, category)
            # For modules with structured spec data (Smart Relay, PLC),
            # also put the cleaned after_ref into specs
            specs_from_after = clean_cid(after_ref)
            specs_from_after = HIMEL_REF_RE.sub('', specs_from_after)
            specs_from_after = re.sub(r'[^\w\s().,/+%-]', ' ', specs_from_after)
            specs_from_after = re.sub(r'\s+', ' ', specs_from_after).strip()
            specs_from_after = re.sub(r'^[^\w(]+|[^\w)]+$', '', specs_from_after).strip()

            if specs_from_after and len(specs_from_after) >= 3:
                specifications = specs_from_after
            else:
                specifications = build_specifications(before_ref, after_ref)

        net_price = apply_discount(list_price)

        products.append({
            'category': category,
            'reference_code': ref_code,
            'product_name': product_name,
            'specifications': specifications,
            'price': net_price,
            'stock_status': STOCK_STATUS,
            'vendor': VENDOR,
            'price_year': DEFAULT_PRICE_YEAR,
        })

    return products


# ── Deduplication ─────────────────────────────────────────────────────────────

def deduplicate(products):
    """Remove duplicate reference codes, keeping first occurrence."""
    seen = set()
    unique = []
    for p in products:
        key = p['reference_code']
        if key not in seen:
            seen.add(key)
            unique.append(p)
    return unique


# ── Output ────────────────────────────────────────────────────────────────────

def save_csv(products, output_path):
    """Save products to CSV."""
    fieldnames = [
        'category', 'reference_code', 'product_name', 'specifications',
        'price', 'stock_status', 'vendor', 'price_year'
    ]
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(products)
    print(f"[OK] CSV saved: {output_path}")


def save_excel(products, output_path):
    """Save products to Excel with basic formatting."""
    if not HAS_EXCEL:
        print("[SKIP] openpyxl not available, skipping Excel export")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Himel Products"

    headers = [
        'category', 'reference_code', 'product_name', 'specifications',
        'price', 'stock_status', 'vendor', 'price_year'
    ]

    # Header row styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, product in enumerate(products, 2):
        for col_idx, field in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=product[field])

    # Column widths
    col_widths = {
        'category': 38, 'reference_code': 26, 'product_name': 52,
        'specifications': 45, 'price': 18, 'stock_status': 14,
        'vendor': 12, 'price_year': 12,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = col_widths.get(header, 20)

    wb.save(output_path)
    print(f"[OK] Excel saved: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    script_dir = Path(__file__).parent
    default_pdf = r'C:\Users\FA\Downloads\PRICE LIST HIMEL 2025 - FORTINDO Disc 45+10_3.pdf'

    pdf_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(default_pdf)
    price_year = int(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_PRICE_YEAR

    if not pdf_path.exists():
        print(f"[ERROR] PDF not found: {pdf_path}")
        sys.exit(1)

    csv_out = script_dir / 'himel_products.csv'
    xlsx_out = script_dir / 'himel_products.xlsx'

    print(f"[INFO] Reading: {pdf_path}")
    print(f"[INFO] Discount: NONE (harga list asli dari PDF, factor = {DISCOUNT_FACTOR:.4f})")

    all_products = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"[INFO] Total pages: {total_pages}\n")

        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if not text:
                print(f"  Page {page_num+1:2d}: (no text, skipped)")
                continue

            # Use first ~500 chars, or full page text if needed for category detection
            header_text = text[:500]
            category = determine_category(header_text)
            if category == 'Other':
                # Try full page text for pages without a clear top header
                category = determine_category(text)

            page_products = extract_products_from_page(page, category)
            print(f"  Page {page_num+1:2d}: [{category}]  →  {len(page_products)} products")
            all_products.extend(page_products)

    print(f"\n[INFO] Total raw rows: {len(all_products)}")
    unique_products = deduplicate(all_products)
    print(f"[INFO] After deduplication: {len(unique_products)} unique products")

    save_csv(unique_products, csv_out)
    save_excel(unique_products, xlsx_out)

    # ── Report ─────────────────────────────────────────────────────────────────
    print(f"\n{'='*65}")
    print(f"EXTRACTION COMPLETE")
    print(f"{'='*65}")
    print(f"Total products  : {len(unique_products)}")
    print(f"Discount factor : {DISCOUNT_FACTOR:.4f}  (NONE — harga list asli)")
    print(f"Vendor          : {VENDOR}")
    print(f"Price year      : {price_year}")

    # Category breakdown
    from collections import Counter
    cats = Counter(p['category'] for p in unique_products)
    print(f"\nCategory breakdown:")
    for cat, count in sorted(cats.items()):
        print(f"  {count:4d}  {cat}")

    # Sample
    print(f"\nSample (first 5 rows):")
    header = f"{'#':<4} {'reference_code':<22} {'product_name':<38} {'price (net)':>15}  category"
    print(header)
    print('-' * len(header))
    for i, p in enumerate(unique_products[:5], 1):
        name = p['product_name'][:36]
        print(f"{i:<4} {p['reference_code']:<22} {name:<38} {p['price']:>15,}  {p['category']}")

    return unique_products


if __name__ == '__main__':
    main()
