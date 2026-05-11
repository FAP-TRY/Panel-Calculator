"""
Extract product catalog from:
  1. CT, Metering, PL, PB, SS, Fuse.pdf
Brand: HOWIG (Indonesian electrical panel accessories manufacturer)
Price list year: 2022 - imported as price_year = 2025 (per task spec)
No discount structure found - prices are direct list prices.
"""

import csv
import re
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import pdfplumber

PDF_PATH = r'C:\Users\FA\Downloads\1. CT, Metering, PL, PB, SS, Fuse.pdf'
CSV_OUT  = r'C:\Projects\Panel Calculator\Tools\ct_metering_pb_fuse_products.csv'
XLSX_OUT = r'C:\Projects\Panel Calculator\Tools\ct_metering_pb_fuse_products.xlsx'

VENDOR = 'HOWIG'
STOCK_STATUS = 1
PRICE_YEAR = 2025

FIELDNAMES = ['category', 'reference_code', 'product_name', 'specifications',
              'price', 'stock_status', 'vendor', 'price_year']


def clean_price(raw):
    """Convert price string like '120.000' or '1.025.000' to integer."""
    if not raw:
        return None
    raw = str(raw).strip()
    if not raw or raw.lower() in ('contact us', '-', 'n/a', ''):
        return None
    raw = raw.replace('Rp', '').replace('IDR', '').strip()
    raw = raw.replace('.', '').replace(',', '').replace(' ', '')
    try:
        return int(float(raw))
    except ValueError:
        return None


def make_row(category, ref_code, product_name, specs, price_raw):
    price = clean_price(price_raw)
    if price is None:
        return None
    ref = str(ref_code).strip() if ref_code else ''
    if not ref:
        return None
    return {
        'category': category,
        'reference_code': ref,
        'product_name': product_name,
        'specifications': specs,
        'price': price,
        'stock_status': STOCK_STATUS,
        'vendor': VENDOR,
        'price_year': PRICE_YEAR,
    }


rows = []

# ─────────────────────────────────────────────────────────────────────────────
# PAGES 1-2: CT WINDOW
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_CT_WINDOW = 'Current Transformer (Window)'

ct_window_data = [
    ('HGS-10',       '10/5A',   '0.5', '2.5',  'Busbar 13mm',  '120.000'),
    ('HGS-20',       '20/5A',   '0.5', '2.5',  'Busbar 13mm',  '120.000'),
    ('HGS-30',       '30/5A',   '0.5', '2.5',  'Busbar 13mm',  '120.000'),
    ('HGS-40',       '40/5A',   '0.5', '2.5',  'Busbar 13mm',  '120.000'),
    ('HG30I-50',     '50/5A',   '1',   '2.5',  '30.5x11mm',    '60.000'),
    ('HG30I-75',     '75/5A',   '1',   '2.5',  '30.5x11mm',    '60.000'),
    ('HG30I-100',    '100/5A',  '1',   '2.5',  '30.5x11mm',    '60.000'),
    ('HG30I-150',    '150/5A',  '1',   '2.5',  '30.5x11mm',    '60.000'),
    ('HG30I-200',    '200/5A',  '1',   '5',    '30.5x11mm',    '60.000'),
    ('HG30I-250',    '250/5A',  '1',   '5',    '30.5x11mm',    '60.000'),
    ('HG30-75',      '75/5A',   '0.5', '2.5',  '31.5x11mm',    '97.000'),
    ('HG30-100',     '100/5A',  '0.5', '2.5',  '31.5x11mm',    '97.000'),
    ('HG30-150',     '150/5A',  '0.5', '5',    '31.5x11mm',    '97.000'),
    ('HG30-200',     '200/5A',  '0.5', '5',    '31.5x11mm',    '97.000'),
    ('HG30-250',     '250/5A',  '0.5', '5',    '31.5x11mm',    '97.000'),
    ('HG40I-300',    '300/5A',  '1',   '5',    '42x11mm',      '64.000'),
    ('HG40I-400',    '400/5A',  '1',   '5',    '42x11mm',      '67.000'),
    ('HG40-300',     '300/5A',  '0.5', '5',    '42x11mm',      '88.000'),
    ('HG40-400',     '400/5A',  '0.5', '5',    '42x11mm',      '91.000'),
    ('HG50I-500',    '500/5A',  '1',   '10',   '51x16mm',      '79.000'),
    ('HG50I-600',    '600/5A',  '1',   '10',   '51x16mm',      '79.000'),
    ('HG50I-800',    '800/5A',  '1',   '10',   '51x16mm',      '101.000'),
    ('HG50-500',     '500/5A',  '0.5', '10',   '51x16mm',      '104.000'),
    ('HG50-600',     '600/5A',  '0.5', '10',   '51x16mm',      '104.000'),
    ('HG52-800',     '800/5A',  '0.5', '10',   '52x31.5mm',    '126.000'),
    ('HG82-800',     '800/5A',  '0.5', '10',   '82x32.5mm',    '170.000'),
    ('HG82-1000',    '1000/5A', '0.5', '10',   '82x32.5mm',    '176.000'),
    ('HG82-1200',    '1200/5A', '0.5', '20',   '82x32.5mm',    '176.000'),
    ('HG82-1250',    '1250/5A', '0.5', '20',   '82x32.5mm',    '176.000'),
    ('HG82-1500',    '1500/5A', '0.5', '20',   '82x32.5mm',    '189.000'),
    ('HG82-1600',    '1600/5A', '0.5', '20',   '82x32.5mm',    '207.000'),
    ('HG82-2000',    '2000/5A', '0.5', '20',   '82x32.5mm',    '222.000'),
    ('HG100-1600',   '1600/5A', '0.5', '20',   '102x32mm',     '231.000'),
    ('HG120-2000',   '2000/5A', '0.5', '20',   '122x52mm',     '262.500'),
    ('HG120-2500',   '2500/5A', '0.5', '40',   '122x52mm',     '288.000'),
    ('HG120-3000',   '3000/5A', '0.5', '40',   '122x52mm',     '305.000'),
    ('HG120-3200',   '3200/5A', '0.5', '40',   '122x52mm',     '325.500'),
    ('HG120-4000',   '4000/5A', '0.5', '40',   '122x52mm',     '420.000'),
    ('HG150-4000',   '4000/5A', '0.5', '40',   '152x134mm',    '602.000'),
    ('HG120-5000',   '5000/5A', '0.5', '40',   '122x52mm',     '473.000'),
    ('HG170-5000',   '5000/5A', '0.5', '40',   '171x103mm',    '825.000'),
    ('HG170-6300',   '6300/5A', '0.5', '40',   '171x103mm',    '925.000'),
    ('HG220-5000',   '5000/5A', '0.5', '30',   '227x65mm',     '925.000'),
    ('HG220-6300',   '6300/5A', '0.5', '30',   '227x65mm',     '1.025.000'),
]

for ref, ratio, cls, burden, hole, price_str in ct_window_data:
    specs = f'Ratio: {ratio}, Class: {cls}, Burden: {burden} VA, Hole: {hole}'
    name = f'CT Window {ref}'
    r = make_row(CATEGORY_CT_WINDOW, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 4: CT SPLIT CORE
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_CT_SPLIT = 'Current Transformer (Split Core)'

ct_split_data = [
    ('HSCT0203-250',  '250/5A',  '1',     '2.5',    '22x33mm',    '450.000'),
    ('HSCT0406-400',  '400/5A',  '1',     '5',      '42x62mm',    '600.000'),
    ('HSCT0406-500',  '500/5A',  '1',     '5',      '42x62mm',    '650.000'),
    ('HSCT0406-600',  '600/5A',  '1',     '5',      '42x62mm',    '675.000'),
    ('HSCT0508-800',  '800/5A',  '0.5/1', '2.5/5',  '52x84mm',    '750.000'),
    ('HSCT0508-1000', '1000/5A', '0.5',   '10',     '52x84mm',    '800.000'),
    ('HSCT0510-1200', '1200/5A', '0.5',   '10',     '52x104mm',   '875.000'),
    ('HSCT0510-1500', '1500/5A', '0.5',   '10',     '52x104mm',   '925.000'),
    ('HSCT0510-1600', '1600/5A', '0.5',   '10',     '52x104mm',   '950.000'),
    ('HSCT0614-2000', '2000/5A', '0.5',   '10',     '63.5x144mm', '1.200.000'),
    ('HSCT0614-2500', '2500/5A', '0.5',   '20',     '63.5x144mm', '1.300.000'),
    ('HSCT0614-3000', '3000/5A', '0.5',   '20',     '63.5x144mm', '1.450.000'),
    ('HSCT0614-3200', '3200/5A', '0.5',   '20',     '63.5x144mm', '1.475.000'),
    ('HSCT0614-4000', '4000/5A', '0.5',   '20',     '63.5x144mm', '1.700.000'),
    ('HSCT0816-4000', '4000/5A', '0.5',   '20',     '82x164mm',   '2.000.000'),
    ('HSCT0816-5000', '5000/5A', '0.5',   '20',     '82x164mm',   '2.300.000'),
    ('HSCT0816-6300', '6300/5A', '0.5',   '20',     '82x164mm',   '2.800.000'),
]

for ref, ratio, cls, burden, hole, price_str in ct_split_data:
    specs = f'Ratio: {ratio}, Class: {cls}, Burden: {burden} VA, Hole: {hole}'
    name = f'CT Split Core {ref}'
    r = make_row(CATEGORY_CT_SPLIT, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 6: CT RING PROTECTION
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_CT_RING_PROT = 'Current Transformer (Ring Protection)'

ct_ring_prot_data = [
    ('HGPR5P10-100',   '100/5A',  '5P10', '5',  '45', '106', '98',  '630.000'),
    ('HGPR5P10-150',   '150/5A',  '5P10', '5',  '45', '106', '83',  '630.000'),
    ('HGPR5P10-200',   '200/5A',  '5P10', '5',  '45', '106', '83',  '630.000'),
    ('HGPR5P10-250',   '250/5A',  '5P10', '5',  '45', '106', '83',  '630.000'),
    ('HGPR5P10-300',   '300/5A',  '5P10', '10', '45', '106', '83',  '640.000'),
    ('HGPR5P10-400',   '400/5A',  '5P10', '10', '45', '106', '83',  '685.000'),
    ('HGPR5P10-500',   '500/5A',  '5P10', '10', '60', '115', '68',  '650.000'),
    ('HGPR5P10-600',   '600/5A',  '5P10', '10', '60', '115', '68',  '650.000'),
    ('HGPR5P10-800',   '800/5A',  '5P10', '15', '90', '135', '60',  '675.000'),
    ('HGPR5P10-1000',  '1000/5A', '5P10', '15', '90', '135', '60',  '700.000'),
    ('HGPR5P10-1200',  '1200/5A', '5P10', '15', '90', '135', '50',  '725.000'),
    ('HGPR5P10-1500',  '1500/5A', '5P10', '15', '90', '135', '50',  '750.000'),
    ('HGPR5P10-1600',  '1600/5A', '5P10', '15', '90', '135', '50',  '775.000'),
    ('HGPR5P10-2000',  '2000/5A', '5P10', '15', '130','180', '40',  '865.000'),
    ('HGPR5P10-2500',  '2500/5A', '5P10', '15', '130','180', '40',  '950.000'),
    ('HGPR5P10-3000',  '3000/5A', '5P10', '15', '130','185', '44',  '1.100.000'),
    ('HGPR5P10-3200',  '3200/5A', '5P10', '15', '130','185', '44',  '1.125.000'),
    ('HGPR5P10-4000S', '4000/5A', '5P10', '15', '140','200', '35',  '1.275.000'),
]

for ref, ratio, cls, burden, id_, od, h, price_str in ct_ring_prot_data:
    specs = f'Ratio: {ratio}, Class: {cls}, Burden: {burden} VA, ID: {id_} mm, OD: {od} mm, H: {h} mm'
    name = f'CT Ring Protection {ref}'
    r = make_row(CATEGORY_CT_RING_PROT, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 7: CT RING MEASUREMENT
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_CT_RING_MEAS = 'Current Transformer (Ring Measurement)'

ct_ring_meas_data = [
    ('HGRG05-100',   '100/5A',  '0.5', '5',  '35',  '84',  '48', '345.000'),
    ('HGRG05-150',   '150/5A',  '0.5', '5',  '33',  '72',  '45', '205.000'),
    ('HGRG05-200',   '200/5A',  '0.5', '5',  '35',  '72',  '36', '175.000'),
    ('HGRG05-250',   '250/5A',  '0.5', '5',  '45',  '72',  '36', '175.000'),
    ('HGRG05-300',   '300/5A',  '0.5', '5',  '45',  '72',  '36', '175.000'),
    ('HGRG05-400',   '400/5A',  '0.5', '5',  '45',  '72',  '36', '175.000'),
    ('HGRG05-500',   '500/5A',  '0.5', '10', '65',  '98',  '33', '225.000'),
    ('HGRG05-600',   '600/5A',  '0.5', '10', '65',  '98',  '33', '230.000'),
    ('HGRG05-800',   '800/5A',  '0.5', '15', '65',  '98',  '33', '235.000'),
    ('HGRG05-1000',  '1000/5A', '0.5', '15', '95',  '135', '35', '480.000'),
    ('HGRG05-1200',  '1200/5A', '0.5', '15', '95',  '135', '35', '500.000'),
    ('HGRG05-1500',  '1500/5A', '0.5', '15', '95',  '135', '35', '520.000'),
    ('HGRG05-1600',  '1600/5A', '0.5', '15', '95',  '135', '35', '530.000'),
    ('HGRG05-2000',  '2000/5A', '0.5', '15', '130', '180', '38', '615.000'),
    ('HGRG05-2500',  '2500/5A', '0.5', '15', '130', '180', '38', '660.000'),
    ('HGRG05-3000',  '3000/5A', '0.5', '15', '130', '180', '38', '710.000'),
    ('HGRG05-3200',  '3200/5A', '0.5', '15', '130', '180', '38', '745.000'),
    ('HGRG05-4000S', '4000/5A', '0.5', '15', '140', '190', '40', '870.000'),
    ('HGRG05-5000S', '5000/5A', '0.5', '15', '175', '229', '40', '1.000.000'),
    ('HGRG05-6300S', '6300/5A', '0.5', '15', '190', '244', '40', '1.350.000'),
    ('HGRG05-4000L', '4000/5A', '0.5', '15', '175', '229', '40', '895.000'),
    ('HGRG05-5000L', '5000/5A', '0.5', '15', '230', '283', '38', '1.220.000'),
    ('HGRG05-6300L', '6300/5A', '0.5', '15', '230', '283', '38', '1.425.000'),
]

for ref, ratio, cls, burden, id_, od, h, price_str in ct_ring_meas_data:
    specs = f'Ratio: {ratio}, Class: {cls}, Burden: {burden} VA, ID: {id_} mm, OD: {od} mm, H: {h} mm'
    name = f'CT Ring Measurement {ref}'
    r = make_row(CATEGORY_CT_RING_MEAS, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 8: CT 3in1
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_CT_3IN1 = 'Current Transformer (3in1)'

ct_3in1_data = [
    ('HGTO242-050', '50/5',  '1.5',  '325.000'),
    ('HGTO242-075', '75/5',  '1.5',  '325.000'),
    ('HGTO242-100', '100/5', '1.5',  '325.000'),
    ('HGTO242-120', '120/5', '1.5',  '325.000'),
    ('HGTO242-125', '125/5', '1.5',  '325.000'),
    ('HGTO242-150', '150/5', '2.5',  '325.000'),
    ('HGTO242-160', '160/5', '2.5',  '325.000'),
    ('HGTO248-300', '300/5', '1.5',  '350.000'),
    ('HGTO248-400', '400/5', '3.75', '350.000'),
    ('HGTO248-500', '500/5', '3.75', '375.000'),
    ('HGTO248-630', '630/5', '3.75', '375.000'),
]

for ref, ratio, burden, price_str in ct_3in1_data:
    specs = f'Ratio: {ratio}, Burden: {burden} VA, Class: 1'
    name = f'CT 3in1 {ref}'
    r = make_row(CATEGORY_CT_3IN1, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGES 9-11: ANALOG METER
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_METER = 'Analog Meter'

# Ammeter
ammeter_data = [
    ('HPM-72/CT',     'Via CT',     '...../5A',  'AC Ammeter, 72x72mm, Class 1.5',        '97.000'),
    ('HPM-96/CT',     'Via CT',     '...../5A',  'AC Ammeter, 96x96mm, Class 1.5',        '97.000'),
    ('HPM-72/D25',    'Direct',     '0-25A',     'AC Ammeter Direct, 72x72mm, Class 1.5', '101.000'),
    ('HPM-72/D40',    'Direct',     '0-40A',     'AC Ammeter Direct, 72x72mm, Class 1.5', '115.000'),
    ('HPM-72/D50',    'Direct',     '0-50A',     'AC Ammeter Direct, 72x72mm, Class 1.5', '115.000'),
    ('HPM-96/D10',    'Direct',     '0-10A',     'AC Ammeter Direct, 96x96mm, Class 1.5', '101.000'),
    ('HPM-96/D20',    'Direct',     '0-20A',     'AC Ammeter Direct, 96x96mm, Class 1.5', '101.000'),
    ('HPM-96/D25',    'Direct',     '0-25A',     'AC Ammeter Direct, 96x96mm, Class 1.5', '101.000'),
    ('HPM-96/D40',    'Direct',     '0-40A',     'AC Ammeter Direct, 96x96mm, Class 1.5', '115.000'),
    ('HPM-96/D50',    'Direct',     '0-50A',     'AC Ammeter Direct, 96x96mm, Class 1.5', '115.000'),
    ('HPM-96/D60',    'Direct',     '0-60A',     'AC Ammeter Direct, 96x96mm, Class 1.5', '115.000'),
    ('HPM-72/A20',    'Direct DC',  '0-20A',     'DC Ammeter Direct, 72x72mm, Class 1.5', '150.000'),
]

for ref, conn, current, specs_base, price_str in ammeter_data:
    specs = f'{specs_base}, Connection: {conn}, Range: {current}'
    name = f'Analog Ammeter {ref}'
    r = make_row(CATEGORY_METER, ref, name, specs, price_str)
    if r:
        rows.append(r)

# Voltmeter
voltmeter_data = [
    ('HPM-72/500',   '0-500VAC', 'AC Voltmeter, 72x72mm, Class 1.5',           '98.000'),
    ('HPM-96/500',   '0-500VAC', 'AC Voltmeter, 96x96mm, Class 1.5',           '98.000'),
    ('HPM-72/V30',   '0-30VDC',  'DC Voltmeter, 72x72mm, Class 1.5',           '150.000'),
    ('HPM-96/DV500', '0-500VAC', 'Double Voltmeter, 96x96mm, Class 1.5',       '475.000'),
]

for ref, voltage, specs_base, price_str in voltmeter_data:
    specs = f'{specs_base}, Range: {voltage}'
    name = f'Analog Voltmeter {ref}'
    r = make_row(CATEGORY_METER, ref, name, specs, price_str)
    if r:
        rows.append(r)

# Other meters
other_meters = [
    ('HPM-96/MD',  'Maximum Demand Ammeter HPM-96/MD',
     'Max Demand Ammeter, Via CT .../5A, 96x96mm, Class 3, Response: 8/15/30min', '225.000'),
    ('HPM-96/KW',  'Kilowatt Meter HPM-96/KW',
     'kW Meter, Via CT .../5A, 380-400VAC, 96x96mm, Class 3', '750.000'),
    ('HPM-96/PF',  'Power Factor Meter HPM-96/PF',
     'Cos Phi Meter, Scale: 0.5 Cap-1-0.5 Ind, 96x96mm, Class 3', '275.000'),
    ('HPM-72/PF',  'Power Factor Meter HPM-72/PF',
     'Cos Phi Meter, Scale: 0.5 Cap-1-0.5 Ind, 72x72mm, Class 3', '280.000'),
    ('HPM-72/FP',  'Frequency Meter Pointer HPM-72/FP',
     'Frequency Meter (Pointer), Range: 45-55Hz, 72x72mm, Class 1.5', '235.000'),
    ('HPM-96/FP',  'Frequency Meter Pointer HPM-96/FP',
     'Frequency Meter (Pointer), Range: 45-55Hz, 96x96mm, Class 1.5', '235.000'),
    ('HPM-96/DFP', 'Double Frequency Meter Pointer HPM-96/DFP',
     'Double Frequency Meter (Pointer), Range: 45-55Hz, 96x96mm', '555.000'),
    ('HPM-72/FV',  'Frequency Meter Vibrating HPM-72/FV',
     'Frequency Meter (Vibrating), Range: 45-55Hz, 72x72mm, Class 1.5', '235.000'),
    ('HPM-96/FV',  'Frequency Meter Vibrating HPM-96/FV',
     'Frequency Meter (Vibrating), Range: 45-55Hz, 96x96mm, Class 1.5', '255.000'),
    ('HPM-96/DFV', 'Double Frequency Meter Vibrating HPM-96/DFV',
     'Double Frequency Meter (Vibrating), Range: 47-53Hz, 96x96mm', '450.000'),
    ('HPM-72/HC',  'Hour Meter HPM-72/HC',
     'Hour Meter, 220/230VAC, 72x72mm, Range: 0-99999.99h', '175.000'),
]

for ref, name, specs, price_str in other_meters:
    r = make_row(CATEGORY_METER, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 12: PILOT LAMP
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_PILOT_LAMP = 'Pilot Lamp'

# LED Pilot Light 22mm - each voltage variant is a separate product with its own ref code
pl22_data = [
    # (ref_230VAC,   ref_110VAC,   ref_24VDC,    color,    price)
    ('HPL-43021', 'HPL-41121', 'HPL-42421', 'Green',  '11.500'),
    ('HPL-43022', 'HPL-41122', 'HPL-42422', 'Red',    '11.500'),
    ('HPL-43023', 'HPL-41123', 'HPL-42423', 'Orange', '11.500'),
    ('HPL-43026', None,        None,        'Yellow', '11.500'),
    ('HPL-43024', None,        None,        'Blue',   '12.500'),
    ('HPL-43025', None,        None,        'White',  '12.500'),
]

for ref_230, ref_110, ref_24, color, price_str in pl22_data:
    for ref, voltage in [(ref_230, '230VAC'), (ref_110, '110VAC'), (ref_24, '24VDC')]:
        if ref and ref != '-':
            specs = f'LED Pilot Light, Size: 22mm, Color: {color}, Voltage: {voltage}'
            name = f'Pilot Lamp 22mm {color} {voltage}'
            r = make_row(CATEGORY_PILOT_LAMP, ref, name, specs, price_str)
            if r:
                rows.append(r)

# LED Pilot Light 16mm
pl16_data = [
    ('HPL-41621', 'Green',  '230VAC', '11.700'),
    ('HPL-41622', 'Red',    '230VAC', '11.700'),
    ('HPL-41623', 'Orange', '230VAC', '11.700'),
]
for ref, color, voltage, price_str in pl16_data:
    specs = f'LED Pilot Light, Size: 16mm, Color: {color}, Voltage: {voltage}'
    name = f'Pilot Lamp 16mm {color} {voltage}'
    r = make_row(CATEGORY_PILOT_LAMP, ref, name, specs, price_str)
    if r:
        rows.append(r)

# LED Pilot Light 30mm
pl30_data = [
    ('HPL-43821', 'Green',  '380VAC', '31.000'),
    ('HPL-43822', 'Red',    '380VAC', '31.000'),
    ('HPL-43823', 'Orange', '380VAC', '31.000'),
]
for ref, color, voltage, price_str in pl30_data:
    specs = f'LED Pilot Light, Size: 30mm, Color: {color}, Voltage: {voltage}'
    name = f'Pilot Lamp 30mm {color} {voltage}'
    r = make_row(CATEGORY_PILOT_LAMP, ref, name, specs, price_str)
    if r:
        rows.append(r)

# LED Flash Buzzer
r = make_row(CATEGORY_PILOT_LAMP, 'HLB-4319', 'LED Flash Buzzer HLB-4319',
             'LED Flash Buzzer, Color: Red, Voltage: 220VAC', '46.500')
if r:
    rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGES 12-13: PUSH BUTTON
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_PB = 'Push Button'

# Monoblock Push Button
mono_pb_data = [
    ('HPB-4102', 'Black',  '1 NO',       '17.000'),
    ('HPB-4107', 'Red',    '1 NC',       '17.000'),
    ('HPB-4111', 'Green',  '1 NO',       '17.000'),
    ('HPB-4108', 'Red',    '1 NO + 1 NC','24.000'),
]
for ref, color, contact, price_str in mono_pb_data:
    specs = f'Monoblock Push Button, Color: {color}, Contact: {contact}, Size: 22mm'
    name = f'Push Button Monoblock {color} {contact}'
    r = make_row(CATEGORY_PB, ref, name, specs, price_str)
    if r:
        rows.append(r)

# Modular Push Button
mod_pb_data = [
    ('HPB-4356', 'Black', '1 NO', '19.000'),
    ('HPB-4358', 'Red',   '1 NC', '19.000'),
    ('HPB-4357', 'Green', '1 NO', '19.000'),
]
for ref, color, contact, price_str in mod_pb_data:
    specs = f'Modular Push Button, Color: {color}, Contact: {contact}'
    name = f'Push Button Modular {color} {contact}'
    r = make_row(CATEGORY_PB, ref, name, specs, price_str)
    if r:
        rows.append(r)

# Modular Illuminated Push Button
illum_pb_data = [
    ('HIPB-R220', 'Red',    '220V',   '1 NC', '43.500'),
    ('HIPB-G220', 'Green',  '220V',   '1 NO', '43.500'),
    ('HIPB-Y220', 'Yellow', '220V',   '1 NO', '43.500'),
    ('HIPB-R024', 'Red',    '24VDC',  '1 NC', '43.500'),
    ('HIPB-G024', 'Green',  '24VDC',  '1 NO', '43.500'),
    ('HIPB-Y024', 'Yellow', '24VDC',  '1 NO', '43.500'),
]
for ref, color, voltage, contact, price_str in illum_pb_data:
    specs = f'Illuminated Push Button, Color: {color}, Voltage: {voltage}, Contact: {contact}'
    name = f'Push Button Illuminated {color} {voltage}'
    r = make_row(CATEGORY_PB, ref, name, specs, price_str)
    if r:
        rows.append(r)

# Double Head Push Button with LED
r = make_row(CATEGORY_PB, 'HDPB-4333', 'Double Head Push Button with LED HDPB-4333',
             'Double Head PB + LED, Color: White, Contact: 1 NC + 1 NO, Voltage: 220VAC', '64.000')
if r:
    rows.append(r)

# Emergency Stop
r = make_row(CATEGORY_PB, 'HEM-4000', 'Emergency Stop Mushroom HEM-4000',
             'Emergency Stop Mushroom, Color: Red, Contact: 1 NC', '30.500')
if r:
    rows.append(r)

# Emergency Stop Cover
r = make_row(CATEGORY_PB, 'HEM-4000PC', 'Protection Cover Emergency Stop HEM-4000 PC',
             'Protection Cover for HEM-4000, Color: Yellow', '19.500')
if r:
    rows.append(r)

# Auxiliary Contacts
aux_data = [
    ('HPB-4396', '1 NO', '7.800'),
    ('HPB-4397', '1 NC', '7.800'),
]
for ref, contact, price_str in aux_data:
    specs = f'Auxiliary Contact, Type: {contact}, For: Modular PB & Selector Switch'
    name = f'Auxiliary Contact {contact}'
    r = make_row(CATEGORY_PB, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGES 13-14: SELECTOR SWITCH
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_SS = 'Selector Switch'

# Monoblock Selector Switch
mono_ss_data = [
    ('HS2P-4141', '2 Position', '1 NO',       '22.500'),
    ('HS3P-4161', '3 Position', '1 NO + 1 NO', '31.500'),
]
for ref, pos, contact, price_str in mono_ss_data:
    specs = f'Monoblock Selector Switch, Position: {pos}, Contact: {contact}, Knob: Black'
    name = f'Selector Switch Monoblock {pos}'
    r = make_row(CATEGORY_SS, ref, name, specs, price_str)
    if r:
        rows.append(r)

# Modular Selector Switch
mod_ss_data = [
    ('HS2P-4363', '2 Position', '1 NO',        '23.500'),
    ('HS3P-4376', '3 Position', '1 NO + 1 NO',  '30.000'),
]
for ref, pos, contact, price_str in mod_ss_data:
    specs = f'Modular Selector Switch, Position: {pos}, Contact: {contact}'
    name = f'Selector Switch Modular {pos}'
    r = make_row(CATEGORY_SS, ref, name, specs, price_str)
    if r:
        rows.append(r)

# Voltmeter Selector Switch
r = make_row(CATEGORY_SS, 'HSV-48', 'Voltmeter Selector Switch HSV-48',
             'Voltmeter SS, Legend: RS-ST-TR-OFF-RN-SN-TN', '105.000')
if r:
    rows.append(r)

# Ammeter Selector Switch
r = make_row(CATEGORY_SS, 'HSA-48', 'Ammeter Selector Switch HSA-48',
             'Ammeter SS, Legend: O-R-S-T', '105.000')
if r:
    rows.append(r)

# OFF-ON Selector Switch
r = make_row(CATEGORY_SS, 'HSOO-48', 'OFF-ON Selector Switch HSOO-48',
             'OFF-ON Selector Switch, Pole: 2', '71.500')
if r:
    rows.append(r)

# MOA/AOM Selector Switch
moa_ss_data = [
    ('HS-1AM', '1 Pole', '65.000'),
    ('HS-2AM', '2 Pole', '87.500'),
    ('HS-3AM', '3 Pole', '107.000'),
]
for ref, pole, price_str in moa_ss_data:
    specs = f'MOA/AOM Selector Switch, {pole}'
    name = f'MOA AOM Selector Switch {pole}'
    r = make_row(CATEGORY_SS, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 14: DIN RAIL
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_DINRAIL = 'DIN Rail'

din_rail_data = [
    ('DRAIL-121F', '1.2 mm', '1000 mm', '1 Function', '21.000'),
    ('DRAIL-122F', '1.2 mm', '1000 mm', '2 Function', '25.000'),
]
for ref, thick, length, func, price_str in din_rail_data:
    specs = f'Aluminium DIN Rail, Thickness: {thick}, Length: {length}, {func}'
    name = f'DIN Rail Aluminium {func}'
    r = make_row(CATEGORY_DINRAIL, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 15: TIME SWITCH
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_TIMESWITCH = 'Time Switch'

time_switch_data = [
    ('HG-PTS16', 'Digital Time Switch HG-PTS16',
     'Digital Time Switch, 1ch, 24h program, 96 segments, 16A-230V, DIN 52.5mm, 150h battery reserve', '290.000'),
    ('HG-ATS16', 'Analog Time Switch HG-ATS16',
     'Analog Time Switch, Daily/Weekly, 16A-230V, DIN 35mm, Calendar 2018-2048, Lithium battery', '265.000'),
]
for ref, name, specs, price_str in time_switch_data:
    r = make_row(CATEGORY_TIMESWITCH, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 15: FUSE HOLDER
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_FUSE_HOLDER = 'Fuse Holder'

fuse_holder_data = [
    ('FH32-S',  '32A', '10x38 mm', 'Without Indicator', '21.800'),
    ('FH32-SL', '32A', '10x38 mm', 'With Indicator',    '23.800'),
    ('FH41-S',  '50A', '14x51 mm', 'Without Indicator', '49.000'),
]
for ref, amp, size, indicator, price_str in fuse_holder_data:
    specs = f'Fuse Holder, Max: {amp}, Size: {size}, Indicator: {indicator}'
    name = f'Fuse Holder {ref} {amp}'
    r = make_row(CATEGORY_FUSE_HOLDER, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 15: FUSE LINK
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_FUSE_LINK = 'Fuse Link'

fuse_link_data = [
    ('FO3802', '2A',  'gG', '10x38 mm', '3.600'),
    ('FO3804', '4A',  'gG', '10x38 mm', '3.600'),
    ('FO3806', '6A',  'gG', '10x38 mm', '3.600'),
    ('FO3810', '10A', 'gG', '10x38 mm', '3.600'),
    ('FO3816', '16A', 'gG', '10x38 mm', '3.600'),
    ('FO3820', '20A', 'gG', '10x38 mm', '3.600'),
    ('FO3825', '25A', 'gG', '10x38 mm', '3.600'),
    ('FO3832', '32A', 'gG', '10x38 mm', '3.600'),
    ('F1440',  '40A', 'gG', '14x51 mm', '9.000'),
    ('F1450',  '50A', 'gG', '14x51 mm', '9.000'),
]
for ref, rating, fuse_type, size, price_str in fuse_link_data:
    specs = f'Fuse Link, Rating: {rating}, Type: {fuse_type}, Size: {size}'
    name = f'Fuse Link {rating} {fuse_type} {size}'
    r = make_row(CATEGORY_FUSE_LINK, ref, name, specs, price_str)
    if r:
        rows.append(r)

# ─────────────────────────────────────────────────────────────────────────────
# WRITE OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

from collections import Counter

print(f'Total rows extracted: {len(rows)}')
print()

cat_counts = Counter(r['category'] for r in rows)
print('Breakdown by category:')
for cat, cnt in sorted(cat_counts.items()):
    print(f'  {cat}: {cnt}')
print()

vendor_counts = Counter(r['vendor'] for r in rows)
print('Breakdown by vendor:')
for v, cnt in sorted(vendor_counts.items()):
    print(f'  {v}: {cnt}')
print()

print('Discount factor applied: None')
print('  All prices are direct list prices from HOWIG 2022 Price List.')
print('  No discount/multiplier structure found in the PDF.')
print()

print('Sample of 5 rows:')
for r in rows[:5]:
    print(f"  ref={r['reference_code']!r}, cat={r['category']!r}, price={r['price']}, specs={r['specifications'][:60]!r}")
print()

# Write CSV
with open(CSV_OUT, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
    writer.writeheader()
    writer.writerows(rows)
print(f'CSV saved: {CSV_OUT}')

# Write XLSX
if HAS_OPENPYXL:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    # Header row with styling
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')
    ws.append(FIELDNAMES)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    # Data rows
    for r in rows:
        ws.append([r[f] for f in FIELDNAMES])
    # Auto column width
    col_widths = {f: len(f) for f in FIELDNAMES}
    for r in rows:
        for f in FIELDNAMES:
            col_widths[f] = max(col_widths[f], len(str(r[f])))
    for i, f in enumerate(FIELDNAMES, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = min(col_widths[f] + 3, 70)
    wb.save(XLSX_OUT)
    print(f'XLSX saved: {XLSX_OUT}')
else:
    print('openpyxl not available - XLSX skipped')

print()
print('Done.')
