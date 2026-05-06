#!/usr/bin/env python3
"""
schneider_pdf_parser.py
=======================
Membaca PDF harga Schneider Electric dan mengekspor ke CSV + Excel
dengan format yang sesuai untuk import ke Kalkulator Panel Tritunggal Swarna.

Output columns:
  category, reference_code, product_name, specifications,
  price, stock_status, vendor, price_year

Stock Status: 1 = Stock, 2 = Indent

Usage:
  python schneider_pdf_parser.py <path_ke_pdf> [tahun_harga]

  Contoh:
  python schneider_pdf_parser.py "Final distribution product schneider.pdf" 2025
"""

import re
import sys
import csv
import os
from pathlib import Path
from datetime import datetime

# Force UTF-8 output on Windows to handle Unicode chars from PDF content
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Dependency check ──────────────────────────────────────────────────────────
try:
    import pdfplumber
except ImportError:
    print("[INFO] Menginstall pdfplumber...")
    os.system("pip install pdfplumber")
    import pdfplumber

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# ── Regex ─────────────────────────────────────────────────────────────────────

# Reference code: starts 1-5 uppercase, then 5+ alphanumeric, total ≥ 6 chars
# Covers: DOMF01102, A9K14106, DOMH12104F, EZ9XPH112, A9C22711, A9A27062, etc.
REF_RE = re.compile(r'\b([A-Z][A-Z0-9]{5,})\b')

# Price: Indonesian thousand-separator format  e.g. 120.500 or 1.234.567
PRICE_RE = re.compile(r'\b(\d{1,3}(?:\.\d{3})+)\b')

# Poles / kutub context lines  e.g. "1 Kutub", "2 Kutub", "1P+N", "3P+N"
POLES_RE = re.compile(
    r'\b(\d\s*Kutub|\d\s*kutub|1P\s*\+\s*N|2P\s*\+\s*N|3P\s*\+\s*N|\d\s*P(?:\+N)?)\b',
    re.IGNORECASE
)

# Full product row: [spec_prefix] REFCODE PRICE SS
# Spec prefix can be up to 100 chars (busbar rows can be long)
PROD_RE = re.compile(
    r'^(.{0,100}?)'               # optional spec prefix (lazy)
    r'\b([A-Z][A-Z0-9]{5,})\b'   # reference code
    r'\s+(\d{1,3}(?:\.\d{3})+)'  # price Indonesian format
    r'\s+([12])\s*$'              # stock status 1 or 2
)

# ── Known false-positive codes ────────────────────────────────────────────────
FALSE_CODES = {
    'HDCFD', 'DOMAE', 'RCCB', 'RCBO', 'ELCB', 'MCB', 'SNI', 'IEC',
    'EN', 'VDC', 'VAC', 'PPN', 'SS', 'IP', 'DC', 'DIN', 'RCP', 'RCU',
    'RCI', 'PDF', 'USB', 'LED', 'LCD', 'CPU', 'RAM', 'ROM',
}

# ── Category keyword map (order = priority, most specific first) ──────────────
CATEGORY_MAP = [
    (['RCBO', 'iDPN N Vigi', 'iDPN', 'RCBO Slim'], 'RCBO'),
    (['RCCB', 'ELCB', 'arus bocor', 'Pengaman Arus Bocor'], 'RCCB/ELCB'),
    (['MCB Box', 'Box Distribusi', 'Distribution Board',
      'DOMH', 'Pintu Transparan', 'IP 40', 'Inbow'], 'MCB Box'),
    (['Busbar Sisir', 'Linergy', 'EZ9XPH', 'Connection Device'], 'Busbar'),
    (['Surge Arrester', 'SPD', 'Penyalur Petir', 'iQuick PRD', 'iPRD', 'PRD'], 'Surge Arrester'),
    (['Kontaktor', 'Contactor', 'iCT', 'Command Control'], 'Kontaktor'),
    (['iACTs', 'Auxilary', 'Auxiliary', 'Spacer kontaktor'], 'Aksesoris Kontaktor'),
    (['Pengaman Peralatan', 'Relay Kontrol', 'RCP', 'RCU', 'RCI'], 'Relay Kontrol'),
    (['Terminal', 'Connector', 'Connection', 'Kabel',
      'Control Switches', 'Disconnection'], 'Connection'),
    (['MCB', 'Miniature Circuit Breaker', 'Domae', 'Easy9',
      'Acti9', 'Acti 9', 'iK60', 'iC60', 'C60', 'C120', 'iCV40', 'ARC', 'AFDD'], 'MCB'),
]

# Lines that are definitely table-header noise to skip
NOISE_STARTS = (
    'Harga (Rp)', 'Sebelum PPN', 'Jumlah kutub', 'Jumlah Kutub',
    'Pengenal (A)', 'Referensi', 'SS (Status', 'Harga belum',
    'Status Stock', '1 (Stock)', '2 (Indent)', 'Sensitivitas',
    'Deskripsi', 'Dimensi (mm)', 'Jumlah Baris', 'Tegangan', 'Waktu',
    'Kontak keluaran', 'Kontak keluran', 'Kontak Tegangan',
    '5-',      # Page number footers like "5-3 Harga belum..."
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def detect_category(text: str) -> str | None:
    tl = text.lower()
    for keywords, cat in CATEGORY_MAP:
        for kw in keywords:
            if kw.lower() in tl:
                return cat
    return None


def is_valid_ref(code: str) -> bool:
    """Check that a matched token looks like a real product reference code."""
    if code in FALSE_CODES:
        return False
    if len(code) < 6:
        return False
    digits = sum(c.isdigit() for c in code)
    letters = sum(c.isalpha() for c in code)
    return digits >= 2 and letters >= 2


def is_noise_line(line: str) -> bool:
    s = line.strip()
    if not s:
        return True
    for prefix in NOISE_STARTS:
        if s.startswith(prefix):
            return True
    # Lines that are only column-header coords: "T L Tb", "1 3 5", etc.
    if re.match(r'^[\d\s\(\)\-/TLb]+$', s) and len(s) < 20:
        return True
    return False


def price_to_int(price_str: str) -> int:
    """'120.500' → 120500"""
    return int(price_str.replace('.', ''))


def clean_spec_prefix(prefix: str, poles_ctx: str) -> str:
    """
    Build a human-readable spec string from:
    - poles_ctx  : accumulated poles context (e.g. "1 Kutub", "1P+N")
    - prefix     : text before ref code on the same line
    """
    # Remove product-series noise words from prefix
    noise_words = [
        r'\bMCB\b', r'\bRCCB\b', r'\bRCBO\b', r'\bDomae\b', r'\bEasy9\b',
        r'\bActi\s*9\b', r'\biDPN\b', r'\biID\b', r'\biCT\b',
        r'\bKontaktor\b', r'"DIN"', r'\bBusbar\s*Sisir\b', r'\bBusbar\b',
        r'\bSlim\b', r'\bN\s*Vigi\b', r'\bVigi\b', r'\bPintu\s*Transparan\b',
    ]
    for nw in noise_words:
        prefix = re.sub(nw, '', prefix, flags=re.I)

    prefix = re.sub(r'\s+', ' ', prefix).strip().strip(',').strip()

    parts = []
    if poles_ctx:
        parts.append(poles_ctx)

    # Check if prefix already contains poles info → update and de-dup
    pm = POLES_RE.search(prefix)
    if pm:
        poles_in_prefix = pm.group(1)
        remainder = (prefix[:pm.start()] + prefix[pm.end():]).strip()
        parts = [poles_in_prefix]   # override ctx with inline info
        if remainder:
            parts.append(remainder)
    elif prefix:
        parts.append(prefix)

    spec = ', '.join(p for p in parts if p).strip(', ')
    # Remove trailing page-number artifact digits like ", 2" or ", 3 5"
    spec = re.sub(r',?\s*\b\d\b\s*$', '', spec).strip()
    return spec


def detect_product_name(line: str) -> str | None:
    """
    Return a cleaned product-line name if this line looks like a section header.
    Returns None if it's a product row or noise.
    """
    s = line.strip()

    # Reject bullet/description lines (never product section headers)
    BULLET_CHARS = ('•', '◾', '⚬', '◦', '–', '▪', '▸', '▹', '★', '☆')
    if s.startswith(BULLET_CHARS):
        return None
    # Reject lines that look like descriptions or specifications
    if s.startswith(('Untuk ', 'Sesuai ', 'Cocok ', 'Proteksi ', 'Kapasitas ',
                      'Standar ', 'Kurva ', 'Daya ', 'Hanya ', 'Catatan ',
                      'Mengaman', 'Mengaktif', 'Menyalak', 'Lebar ',
                      'Pengaman motor', 'Index Proteksi', 'Tipe Inbow')):
        return None

    # Must not be a product row (has ref+price)
    if PROD_RE.match(s):
        return None
    # Remove page decorations
    s = re.sub(r'\s*HDCFD\s*', '', s)
    s = re.sub(r'■\s*', '', s)
    # Remove trailing page numbers like "2", "5-3"
    s = re.sub(r'\s+\d+\s*$', '', s).strip()
    s = re.sub(r'\s+5-\d+\s*$', '', s).strip()

    if len(s) < 5:
        return None

    # Must mention a known product keyword
    has_kw = any(kw.lower() in s.lower() for kw in [
        'MCB', 'RCCB', 'RCBO', 'Busbar', 'Surge', 'Kontaktor', 'Relay',
        'Easy9', 'Domae', 'Acti9', 'Acti 9', 'iDPN', 'iID', 'iCT', 'Linergy',
        'Pengaman Arus Bocor', 'Pengaman Peralatan', 'iACTs', 'Connection',
        'iK60', 'iC60', 'C60', 'C120', 'iQuick', 'iPRD', 'Box Distribusi',
        'MCB Box', 'iCV40', 'ARC', 'Control Switches',
    ])
    return s if has_kw else None


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_pdf(pdf_path: str, price_year: int) -> list[dict]:
    products = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"[INFO] Total halaman: {total_pages}")
        print(f"[INFO] Tahun harga  : {price_year}")
        print()

        for page_idx, page in enumerate(pdf.pages):
            if page_idx == 0:
                continue  # Skip halaman 1 (daftar isi)

            text = page.extract_text()
            if not text:
                continue

            lines = [ln.strip() for ln in text.split('\n')]
            page_num = page_idx + 1

            # ── Determine page-level category & product name ──────────
            header_block = ' '.join(lines[:8])
            page_category = detect_category(header_block) or 'MCB'
            page_product_name = ''
            for ln in lines[:6]:
                name = detect_product_name(ln)
                if name:
                    page_product_name = name
                    break

            # ── Per-line state ────────────────────────────────────────
            current_category = page_category
            current_product_name = page_product_name
            current_poles = ''          # e.g. "1 Kutub", "1P+N"
            page_products = 0

            for line in lines:
                if is_noise_line(line):
                    continue

                m = PROD_RE.match(line)
                if m:
                    prefix   = m.group(1)
                    ref_code = m.group(2)
                    price_s  = m.group(3)
                    ss       = int(m.group(4))

                    if not is_valid_ref(ref_code):
                        continue

                    spec = clean_spec_prefix(prefix, current_poles)

                    products.append({
                        'category'      : current_category,
                        'reference_code': ref_code,
                        'product_name'  : current_product_name,
                        'specifications': spec,
                        'price'         : price_to_int(price_s),
                        'stock_status'  : ss,
                        'vendor'        : 'Schneider Electric',
                        'price_year'    : price_year,
                    })
                    page_products += 1

                else:
                    # Not a product row — update context

                    # Category update
                    new_cat = detect_category(line)
                    if new_cat:
                        current_category = new_cat

                    # Product name update
                    new_name = detect_product_name(line)
                    if new_name:
                        current_product_name = new_name

                    # Poles context update
                    pm = POLES_RE.search(line)
                    if pm:
                        current_poles = pm.group(1)
                        # Normalize: "1 Kutub" / "2 Kutub" / "1P+N" etc.
                        current_poles = re.sub(r'\s+', ' ', current_poles).strip()

            print(f"  [P{page_num:02d}] {page_products:3d} produk | "
                  f"cat={current_category:20s} | {current_product_name[:45]}")

    return products


# ── Export: CSV ───────────────────────────────────────────────────────────────

FIELDNAMES = [
    'category', 'reference_code', 'product_name', 'specifications',
    'price', 'stock_status', 'vendor', 'price_year',
]

def save_csv(products: list[dict], out_path: Path) -> None:
    with open(out_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(products)
    print(f"\n[OK] CSV  → {out_path}  ({len(products)} baris)")


# ── Export: Excel ─────────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'Category', 'Reference Code', 'Product Name', 'Specifications',
    'Price (Rp)', 'Stock Status', 'Vendor', 'Price Year',
]

def save_excel(products: list[dict], out_path: Path) -> None:
    if not HAS_EXCEL:
        print("[WARN] openpyxl tidak tersedia — skip Excel output.")
        print("       Jalankan: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produk Schneider"

    # ── Header row ────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────
    alt_fill = PatternFill(start_color='f0f9ff', end_color='f0f9ff', fill_type='solid')
    for row_idx, p in enumerate(products, 2):
        values = [
            p['category'], p['reference_code'], p['product_name'],
            p['specifications'], p['price'], p['stock_status'],
            p['vendor'], p['price_year'],
        ]
        use_alt = (row_idx % 2 == 0)
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if use_alt:
                cell.fill = alt_fill
            # Center-align numeric/short columns
            if col in (2, 6, 7, 8):   # ref_code, ss, vendor, year
                cell.alignment = Alignment(horizontal='center')
            if col == 5:               # price
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    # ── Column widths ─────────────────────────────────────────────
    col_widths = [18, 16, 40, 30, 16, 13, 20, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = w

    # ── Freeze header + auto-filter ───────────────────────────────
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)
    print(f"[OK] Excel → {out_path}  ({len(products)} baris)")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    pdf_path = sys.argv[1]
    price_year = int(sys.argv[2]) if len(sys.argv) > 2 else datetime.now().year

    if not os.path.isfile(pdf_path):
        print(f"[ERROR] File tidak ditemukan: {pdf_path}")
        sys.exit(1)

    print("=" * 60)
    print("  Schneider PDF -> CSV/Excel Parser")
    print("=" * 60)
    print(f"  Input : {pdf_path}")
    print(f"  Tahun : {price_year}")
    print("=" * 60)
    print()

    products = parse_pdf(pdf_path, price_year)

    print()
    print("=" * 60)
    print(f"  Total produk ditemukan: {len(products)}")
    print("=" * 60)

    if not products:
        print("\n[ERROR] Tidak ada produk yang berhasil di-parse.")
        print("        Pastikan PDF adalah versi yang benar.")
        sys.exit(1)

    # ── Show sample ───────────────────────────────────────────────
    print("\nSample 8 produk pertama:")
    print(f"  {'Category':<18} {'RefCode':<12} {'ProductName':<30} {'Spec':<22} {'Price':>10}  SS")
    print("  " + "-" * 100)
    for p in products[:8]:
        print(
            f"  {p['category']:<18} {p['reference_code']:<12} "
            f"{p['product_name'][:28]:<30} {p['specifications'][:20]:<22} "
            f"{p['price']:>10,}   {p['stock_status']}"
        )

    # ── Count by category ─────────────────────────────────────────
    from collections import Counter
    cat_counts = Counter(p['category'] for p in products)
    print("\nJumlah produk per kategori:")
    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f"  {cat:<25} : {cnt}")

    # ── Save outputs ──────────────────────────────────────────────
    base = Path(pdf_path).stem
    out_dir = Path(pdf_path).parent
    csv_path  = out_dir / f"{base}_produk.csv"
    xlsx_path = out_dir / f"{base}_produk.xlsx"

    print()
    save_csv(products, csv_path)
    save_excel(products, xlsx_path)

    print()
    print("Selesai! File tersimpan di:")
    print(f"  CSV  : {csv_path}")
    print(f"  Excel: {xlsx_path}")
    print()
    print("Import ke aplikasi via menu Pengaturan → Import Produk dari CSV/Excel")


if __name__ == '__main__':
    main()
