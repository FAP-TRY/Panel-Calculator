import pdfplumber
import csv
import re
from collections import Counter

pdf_path = r'C:\Users\FA\Downloads\2. NH Fuse, LBS, Isolator, Digital Meter, kWh.pdf'
out_csv  = r'C:\Projects\Panel Calculator\Tools\nh_fuse_lbs_meter_products.csv'
out_xlsx = r'C:\Projects\Panel Calculator\Tools\nh_fuse_lbs_meter_products.xlsx'

def clean_price(s):
    if not s or str(s).strip() in ('', '-', 'Contact Us', 'None'):
        return None
    s = re.sub(r'[^0-9]', '', str(s))
    return int(s) if s else None

def clean_code(s):
    return str(s).strip() if s else None

rows = []

with pdfplumber.open(pdf_path) as pdf:
    full_texts = [p.extract_text() or '' for p in pdf.pages]

# ── PAGE 1: NH Fuse gG/gL ──────────────────────────────────────────────────
page1_text = full_texts[0]
for line in page1_text.split('\n'):
    m = re.match(r'^(0\d{4})\s+(\d+)\s+500~\s+120\s+([\d.]+)$', line.strip())
    if m:
        code, amps, price_str = m.group(1), m.group(2), m.group(3)
        price = clean_price(price_str)
        code_int = int(code)
        if code_int <= 5860:
            size = '00'
        elif code_int <= 5893:
            size = '1'
        elif code_int <= 5884:
            size = '2'
        else:
            size = '3'
        rows.append({
            'category': 'NH Fuse',
            'reference_code': code,
            'product_name': 'NH Fuse gG/gL Size ' + size,
            'specifications': amps + 'A; 500VAC; Breaking Capacity 120kA',
            'price': price,
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 1 (NH Fuse):', len(rows))

# ── PAGE 2: NH Fuse Base Ceramic ───────────────────────────────────────────
page2_text = full_texts[1]
for line in page2_text.split('\n'):
    m = re.match(r'^(HFT\S+)\s+(\S+)\s+(\d+)\s+([\d~]+)\s+([\d.,]+)$', line.strip())
    if m:
        code, size, max_amp, volt, price_str = m.groups()
        price = clean_price(price_str)
        if price:
            rows.append({
                'category': 'NH Fuse Base',
                'reference_code': clean_code(code),
                'product_name': 'NH Fuse Base (Ceramic Insulation)',
                'specifications': 'Size ' + size + '; Max ' + max_amp + 'A; ' + volt + 'VAC',
                'price': price,
                'stock_status': 1,
                'vendor': 'HOWIG',
                'price_year': 2025,
            })
print('After page 2 (NH Fuse Base Ceramic):', len(rows))

# ── PAGE 3: NH Fuse Base DMC + Fuse Puller ────────────────────────────────
page3_text = full_texts[2]
for line in page3_text.split('\n'):
    m = re.match(r'^(HFT\S+)\s+(\S+)\s+(\d+)\s+([\d~]+)\s+([\d.,]+)$', line.strip())
    if m:
        code, size, max_amp, volt, price_str = m.groups()
        price = clean_price(price_str)
        if price:
            rows.append({
                'category': 'NH Fuse Base',
                'reference_code': clean_code(code),
                'product_name': 'NH Fuse Base (DMC)',
                'specifications': 'Size ' + size + '; Max ' + max_amp + 'A; ' + volt + 'VAC',
                'price': price,
                'stock_status': 1,
                'vendor': 'HOWIG',
                'price_year': 2025,
            })

m2 = re.search(r'(HGFH\s*-\s*123)\s+([\d,\s&]+)\s+1000V\s+([\d.]+)', page3_text)
if m2:
    rows.append({
        'category': 'NH Fuse Accessories',
        'reference_code': m2.group(1).replace(' ', ''),
        'product_name': 'Fuse Puller',
        'specifications': 'For NH Fuse Size ' + m2.group(2).strip() + '; 1000V Insulation',
        'price': clean_price(m2.group(3)),
        'stock_status': 1,
        'vendor': 'HOWIG',
        'price_year': 2025,
    })
print('After page 3 (NH Fuse Base DMC):', len(rows))

# ── PAGE 4: Horizontal Fuse Switch + Surge Protection Device ───────────────
page4_text = full_texts[3]
for line in page4_text.split('\n'):
    m = re.match(r'^(HFS\s+\S+)\s+(\d+)A\s+[\d.]+A\s+\d+kA\s+\d+kA\s+([\d.,]+)$', line.strip())
    if m:
        code, max_amp, price_str = m.groups()
        price = clean_price(price_str)
        rows.append({
            'category': 'Horizontal Fuse Switch',
            'reference_code': clean_code(code),
            'product_name': 'Horizontal Fuse Switch (HFS)',
            'specifications': 'Max ' + max_amp + 'A 400VAC; 100kA breaking capacity',
            'price': price,
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })

for line in page4_text.split('\n'):
    m = re.match(r'^(BY7-\S+)\s+(\S+)\s+(\w+)\s+(\w+)\s+(\d+)\s+([\d/]+)\s+(\d+)\s+([\d.]+)\s+([\d.,]+)$', line.strip())
    if m:
        code = m.group(1)
        pole = m.group(2)
        model = m.group(3) + ' ' + m.group(4)
        imax = m.group(5)
        un = m.group(6)
        uc = m.group(7)
        price_str = m.group(9)
        price = clean_price(price_str)
        rows.append({
            'category': 'Surge Protection Device',
            'reference_code': clean_code(code),
            'product_name': 'Surge Protection Device (SPD)',
            'specifications': pole + '; ' + model + '; Imax ' + imax + 'kA; Un ' + un + 'V; Uc ' + uc + 'V',
            'price': price,
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 4 (HFS + SPD):', len(rows))

# ── PAGE 5: Cylindrical / Octagonal-Hex / Drum Insulators ────────────────
page5_text = full_texts[4]
for line in page5_text.split('\n'):
    m = re.match(r'^(HBCI\s*-\s*\S+)\s+([\d./]+)\s+(\d+)\s+(\S+)\s+([\d.]+)$', line.strip())
    if m:
        code, diam, height, bolt, price_str = m.groups()
        rows.append({
            'category': 'Insulator',
            'reference_code': clean_code(code),
            'product_name': 'Cylindrical Insulator (HBCI)',
            'specifications': 'D1/D2 ' + diam + 'mm; H ' + height + 'mm; Bolt ' + bolt + '; DMC',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })

for line in page5_text.split('\n'):
    m = re.match(r'^((?:HBOC|HBHEX)\s*-\s*\S+)\s+([\d,./]+)\s+(\d+)\s+(\S+)\s+([\d.]+)$', line.strip())
    if m:
        code, diam, height, bolt, price_str = m.groups()
        rows.append({
            'category': 'Insulator',
            'reference_code': clean_code(code),
            'product_name': 'Octagonal/Hexagonal Insulator',
            'specifications': 'D1/D2 ' + diam + 'mm; H ' + height + 'mm; Bolt ' + bolt + '; DMC',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })

for line in page5_text.split('\n'):
    m = re.match(r'^(HBSM\s*-\s*\S+)\s+([\d./]+)\s+(\d+)\s+(\S+)\s+([\d.]+)$', line.strip())
    if m:
        code, diam, height, bolt, price_str = m.groups()
        rows.append({
            'category': 'Insulator',
            'reference_code': clean_code(code),
            'product_name': 'Drum Insulator (HBSM)',
            'specifications': 'D1/D2 ' + diam + 'mm; H ' + height + 'mm; Bolt ' + bolt + '; DMC',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 5 (Insulators):', len(rows))

# ── PAGE 6: Step Insulator + MV Insulator ─────────────────────────────────
page6_text = full_texts[5]
for line in page6_text.split('\n'):
    m = re.match(r'^(HBSI\s*-\s*\S+)\s+(\d+)\s+(M\d+\s*\(x\d+\))\s+([\d.]+)$', line.strip())
    if m:
        code, busbar, bolt, price_str = m.groups()
        rows.append({
            'category': 'Insulator',
            'reference_code': clean_code(code),
            'product_name': 'Step Insulator (HBSI)',
            'specifications': 'Busbar ' + busbar + 'mm; Bolt ' + bolt + '; DMC',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })

for line in page6_text.split('\n'):
    m = re.match(r'^(EL-\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+\S+.*?([\d.]+)$', line.strip())
    if m:
        code, kv, diam, height, price_str = m.groups()
        rows.append({
            'category': 'Insulator',
            'reference_code': clean_code(code),
            'product_name': 'MV Insulator (EL Series)',
            'specifications': kv + 'kV; D ' + diam + 'mm; H ' + height + 'mm; DMC',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 6 (Step + MV Insulators):', len(rows))

# ── PAGE 7: Busbar Support (codes include spaces like "HBS1P100 - 306/211") ──
page7_text = full_texts[6]
for line in page7_text.split('\n'):
    # Match: HBSxP... <length> <busbar> <price>
    m = re.match(r'^(HBS(\dP)[^\t]+?)\s{2,}(\d+)\s+([\dx/\s]+?)\s+([\d.]+)$', line.strip())
    if not m:
        m = re.match(r'^(HBS(\dP)\S+(?:\s+-\s+\S+)?)\s+(\d+)\s+([\dx/ ]+)\s+([\d.]+)$', line.strip())
    if m:
        code, poles_tag, length, busbar, price_str = m.groups()
        poles = poles_tag[0] + ' Pole'
        rows.append({
            'category': 'Busbar Support',
            'reference_code': clean_code(code),
            'product_name': 'Busbar Support ' + poles + ' (HBS' + poles_tag + ')',
            'specifications': poles + '; L ' + length + 'mm; Busbar ' + busbar.strip() + 'mm; SMC',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })

# Also add insulating pipe
for line in page7_text.split('\n'):
    m = re.match(r'^(HBIP\S+)\s+(\d+)\s+(\d+)\s+(\d+)\s+([\d.]+)$', line.strip())
    if m:
        code, id_mm, od_mm, length, price_str = m.groups()
        rows.append({
            'category': 'Busbar Support',
            'reference_code': clean_code(code),
            'product_name': 'Insulating Pipe (HBIP)',
            'specifications': 'ID ' + id_mm + 'mm; OD ' + od_mm + 'mm; L ' + length + 'mm; FRP',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 7 (Busbar Support):', len(rows))

# ── PAGE 8: HBEL Busbar Support ───────────────────────────────────────────
page8_text = full_texts[7]
for line in page8_text.split('\n'):
    m = re.match(r'^(HBEL\s*-\s*\S+)\s+(\d+\s+Pole)\s+(\d+)\s+([\dx/\s]+)\s+([\d.]+)$', line.strip())
    if m:
        code, poles, length, busbar, price_str = m.groups()
        rows.append({
            'category': 'Busbar Support',
            'reference_code': clean_code(code),
            'product_name': 'Busbar Support (HBEL)',
            'specifications': poles + '; L ' + length + 'mm; Busbar ' + busbar.strip() + 'mm',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 8 (HBEL Busbar Support):', len(rows))

# ── PAGE 9: Power Supply ───────────────────────────────────────────────────
page9_text = full_texts[8]
for line in page9_text.split('\n'):
    m = re.match(r'^(HGPS[\s-]\S+)\s+([\d~VAC/]+)\s+(\d+)\s+(\d+)\s+(\S+)\s+([\d.,]+)$', line.strip())
    if m:
        code, vin, iout, vout, dim, price_str = m.groups()
        rows.append({
            'category': 'Power Supply',
            'reference_code': clean_code(code),
            'product_name': 'DIN Rail Power Supply',
            'specifications': 'Input ' + vin + 'VAC; ' + iout + 'A; Output ' + vout + 'VDC; ' + dim + 'mm',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 9 (Power Supply):', len(rows))

# ── PAGE 10: Harmonic Blocking Reactor ────────────────────────────────────
page10_text = full_texts[9]
for line in page10_text.split('\n'):
    m = re.match(r'^(HFT-DR\S+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+(\S+)\s+(\d+)\s+([\d.,]+)$', line.strip())
    if m:
        code = m.group(1)
        kvar = m.group(2)
        cap_kvar = m.group(4)
        price_str = m.group(9)
        rows.append({
            'category': 'Harmonic Blocking Reactor',
            'reference_code': clean_code(code),
            'product_name': 'Detuned Reactor for Capacitor Bank',
            'specifications': kvar + ' kVAr @400VAC; ' + cap_kvar + ' kVAr Capacitor; 525VAC; p=7%; fr=189Hz',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 10 (Reactor):', len(rows))

# ── PAGE 11: Digital Power Monitoring ─────────────────────────────────────
digital_pm_data = [
    ('HFPM-96F', 'Via CT /5A or /1A', 'RS485 Modbus', 'Multi', '96x96mm', '65~480VAC', 2000000),
    ('HFPM-96I', 'Via CT /5A or /1A', 'RS485 Modbus/Ethernet', 'Multi', '96x96mm', '65~480VAC', 3400000),
]
for code, conn, comm, tariff, size, volt, price in digital_pm_data:
    rows.append({
        'category': 'Digital Power Monitor',
        'reference_code': code,
        'product_name': 'Digital Power Monitoring (HFPM)',
        'specifications': conn + '; ' + comm + '; ' + tariff + ' Tariff; ' + size + '; ' + volt,
        'price': price,
        'stock_status': 1,
        'vendor': 'HOWIG',
        'price_year': 2025,
    })
print('After page 11 (Power Monitor):', len(rows))

# ── PAGE 12: Digital Meter ─────────────────────────────────────────────────
digital_meter_data = [
    ('HDM9696-1I',   'Single Phase Digital Ammeter',            'Via CT /5A; 96x96mm; 220VAC Aux Power; Class 0.5', 200000),
    ('HDM9696-1U',   'Single Phase Digital Voltmeter',          '0-600VAC; 96x96mm; 220VAC Aux Power; Class 0.5',  200000),
    ('HDM9696-3I',   'Three Phase Digital Ammeter',             'Via CT /5A; 96x96mm; 220VAC Aux Power; Class 0.5', 375000),
    ('HDM9696-3U',   'Three Phase Digital Voltmeter',           '0-600VAC; 96x96mm; 220VAC Aux Power; Class 0.5',  375000),
    ('HDM9696-3UI',  'Three Phase Digital Volt+Ammeter',        '0-600VAC; 96x96mm; 220VAC Aux Power; Class 0.5',  450000),
    ('HDM9696-3UIF', 'Three Phase Digital V+A+Frequency Meter', '0-600VAC; 96x96mm; 220VAC Aux Power; Class 0.5',  500000),
    ('HDM9648-U30',  'DC Voltmeter',                            '0-30VDC; 96x48mm; Direct Connection',             225000),
    ('HDM9648-I30',  'DC Ammeter',                              '0-30ADC; 96x48mm; Via Shunt Resistor',            225000),
    ('HDM-SHUNT',    'Shunt Resistor for DC Ampere Meter',      '0-75mV',                                           55000),
]
for code, name, specs, price in digital_meter_data:
    rows.append({
        'category': 'Digital Meter',
        'reference_code': code,
        'product_name': name,
        'specifications': specs,
        'price': price,
        'stock_status': 1,
        'vendor': 'HOWIG',
        'price_year': 2025,
    })
print('After page 12 (Digital Meter):', len(rows))

# ── PAGE 13: Energy Meter (kWh) ───────────────────────────────────────────
energy_meter_data = [
    ('HFEM-1140MDP',      '1P Analog Energy Meter; Direct Connection 40A; 1 Module',              'Class 1; Pulse Output',          290000),
    ('HFEM-1132DP',       '1P Digital Energy Meter; Direct Connection 32A; 1 Module',             'Class 1; Pulse Output',          315000),
    ('HFEM-1132DCOM',     '1P Digital Multifunction Energy Meter; Direct Connection 32A',         'Class 1; Pulse Output & RS485',  570000),
    ('HFEM-1263MDP',      '1P Analog Energy Meter; Direct Connection 63A; 2 Modules',             'Class 1; Pulse Output',          305000),
    ('HFEM-1263DDP',      '1P Digital Multifunction Energy Meter; Direct Connection 63A; 2M',     'Class 1; Pulse Output',          425000),
    ('HFEM-1263DCOM-MT',  '1P Digital Multifunction; Direct Connection 63A; 4M; Multi Tariff',   'Class 1; Pulse Output & RS485',  780000),
    ('HFEM-1463DPP',      '1P Digital Multifunction; Direct Connection 63A; 4M',                  'Class 0.5s; Pulse Output & RS485', 950000),
    ('HFEM-3400CT-DP',    '3P Digital Energy Meter; Via CT /5A; 4 Modules',                       'Class 1; Pulse Output',          925000),
    ('HFEM-3480DDP',      '3P Digital Energy Meter; Direct Connection 80A; 4 Modules',            'Class 1; Pulse Output',          875000),
    ('HFEM-3780DCOM-MT',  '3P Digital Multimeasurement; Direct 80A; 7M; Multi Tariff',           'Class 1; RS485 & Pulse Output', 1850000),
    ('HFEM-3700CTCOM-MT', '3P Digital Multimeasurement; Via CT /5A; 7M; Multi Tariff',           'Class 1; RS485 & Pulse Output', 1750000),
    ('HFEM-3480DCOM-MT',  '3P Digital Multimeasurement; Direct 80A; 4M; Multi Tariff',           'Class 1; RS485 & Pulse Output', 2100000),
    ('HFEM-3400CTCOM-MT', '3P Digital Multimeasurement; Via CT /5A; 4M; Multi Tariff',           'Class 0.5s; RS485 & Pulse Output', 2000000),
]
for code, desc, specs, price in energy_meter_data:
    rows.append({
        'category': 'kWh Meter',
        'reference_code': code,
        'product_name': 'Energy Meter (HFEM)',
        'specifications': desc + '; ' + specs,
        'price': price,
        'stock_status': 1,
        'vendor': 'HOWIG',
        'price_year': 2025,
    })
print('After page 13 (kWh Meter):', len(rows))

# ── PAGE 14: Load Break Switch ────────────────────────────────────────────
page14_text = full_texts[13]
for line in page14_text.split('\n'):
    m = re.match(r'^(VL-LBS\S+)\s+(Load Break Switch\s+\d+\s+Pole\s+(\d+)A)\s+([\d.,]+)$', line.strip())
    if m:
        code = m.group(1)
        amps = m.group(3)
        price_str = m.group(4)
        rows.append({
            'category': 'LBS',
            'reference_code': clean_code(code),
            'product_name': 'Load Break Switch (LBS) 3 Pole',
            'specifications': amps + 'A; 3 Pole; AC23A Category',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 14 (LBS):', len(rows))

# ── PAGE 15: Manual Changeover Switch ────────────────────────────────────
page15_text = full_texts[14]
for line in page15_text.split('\n'):
    m = re.match(r'^(VL-COS\S+)\s+(Manual Changeover Switch\s+\d+\s+Pole\s+(\d+)A)\s+([\d.,]+)$', line.strip())
    if m:
        code = m.group(1)
        amps = m.group(3)
        price_str = m.group(4)
        rows.append({
            'category': 'Changeover Switch',
            'reference_code': clean_code(code),
            'product_name': 'Manual Changeover Switch (COS) 4 Pole',
            'specifications': amps + 'A; 4 Pole; AC23A Category',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 15 (Manual COS):', len(rows))

# ── PAGE 16: Motorized Changeover Switch (ATS) ───────────────────────────
page16_text = full_texts[15]
for line in page16_text.split('\n'):
    m = re.match(r'^(VL-ATS\S+)\s+(Motorized Changeover Switch\s+\d+\s+Pole\s+(\d+)A)\s+([\d.,]+)$', line.strip())
    if m:
        code = m.group(1)
        amps = m.group(3)
        price_str = m.group(4)
        rows.append({
            'category': 'Motorized Changeover Switch',
            'reference_code': clean_code(code),
            'product_name': 'Motorized Changeover Switch (ATS) 4 Pole',
            'specifications': amps + 'A; 4 Pole; Motor Operated',
            'price': clean_price(price_str),
            'stock_status': 1,
            'vendor': 'HOWIG',
            'price_year': 2025,
        })
print('After page 16 (ATS):', len(rows))

# ── Filter out rows without valid price ───────────────────────────────────
rows = [r for r in rows if r['price'] is not None and r['price'] > 0]

print()
print('Total valid rows:', len(rows))

cat_counts = Counter(r['category'] for r in rows)
print()
print('Breakdown by category:')
for cat, count in sorted(cat_counts.items()):
    print('  ' + cat + ': ' + str(count))

print()
print('Sample 5 rows:')
for r in rows[:5]:
    print(r)

# ── Write CSV ─────────────────────────────────────────────────────────────
fieldnames = ['category','reference_code','product_name','specifications','price','stock_status','vendor','price_year']
with open(out_csv, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
print()
print('CSV saved:', out_csv)

# ── Write XLSX ────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F497D')
        cell.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append([r[f] for f in fieldnames])
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    wb.save(out_xlsx)
    print('XLSX saved:', out_xlsx)
except Exception as e:
    print('XLSX error:', str(e))

print()
print('Discount: None applied. PDF contains direct net prices (no list price + discount structure).')
